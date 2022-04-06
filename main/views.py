import io
import json
import zipfile

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render
from django.utils.translation.trans_null import gettext_lazy as _
from openpyxl import load_workbook
from PIL import Image

from main.models import DiplomaTemplate, ExcelForGenerate, ZipFile
from main.serializers import DiplomaSerializer

from .utils import *


@login_required
def index(request):
    return render(request, 'index.html')


@login_required
def editor(request):
    return render(request, 'diploma_editor.html')


def get_diplomas_templates(request):
    if not request.user.is_authenticated:
        return JsonResponse({'result': False, 'message': _('User is not authenticated.')})

    templates = DiplomaTemplate.objects.all()
    return JsonResponse({
        'result': True,
        'message': _('Templates were returned.'),
        'templates': DiplomaSerializer(templates, many=True).data
    })


def upload_templates(request):
    '''
    View-функция для загрузки шаблона грамоты
    '''
    if request.method == 'POST':
        f = request.FILES
        temp = DiplomaTemplate.objects.create(diploma=f.get('file'))
        try:
            temp.save()
        except Exception as e:
            return JsonResponse({'result': False, 'message': _(e.__str__())})
        return JsonResponse({
            'result': True, 
            'message': _('Template was saved.'), 
            'url': temp.diploma.url, 
            'id': temp.id
        })

    elif request.method == 'DELETE':
        id = request.POST.get('id')
        dp = DiplomaTemplate.objects.get(pk=id)
        dp.delete()
        return JsonResponse({'result': True, 'message': _('Template delete successful.')})

    return JsonResponse({'result': False, 'message': _('Method not allowed.')})


def generate_diploma(request):
    '''
    View-функция для генерации грамот
    '''
    if request.method == 'GET':
        # получяем массив с именами учеников
        names = json.loads(request.GET.get('names'))

        # Если длинна массива равна одному, значит пришел запрос на тестовую грамоту, следовательно zip-файла создавать не нужно
        if len(names) <= 1:
            url = generate_image_object(request.GET, names[0], False)
            return JsonResponse({
                'result': True,
                'message': _('Images was generated.'),
                'url': url
            })
        
        buffer = io.BytesIO()
        zip_file = zipfile.ZipFile(buffer, 'w')

        for i in names:
            b = io.BytesIO()

            path = generate_image_object(request.GET, i, True)
            image = Image.open(path)
            image.save(b, format='JPEG')

            b.seek(0)

            zip_file.writestr(path.split('\\')[-1], b.read())

        zip_file.close()

        f = ZipFile.objects.create(file=InMemoryUploadedFile(buffer, None, "TestZip.zip", 'application/zip', buffer.tell, None))
        f.save()

        return JsonResponse({'result': True, 'message': _('Images has been generated.'), 'url': f.file.url})

def get_names(request):
    if request.method == 'POST':
        names = []

        excel = ExcelForGenerate(file=request.FILES.get('file'))
        try:
            excel.save()
        except Exception as e:
            return JsonResponse({
                'result': False,
                'message': _(e.__str__()),
            })

        wb_obj = load_workbook(excel.file.path)

        sheet_obj = wb_obj.active

        max_row = sheet_obj.max_row

        for i in range(2, max_row + 1):
            cell = sheet_obj.cell(i, 1)
            names.append(cell.value)

        return JsonResponse({
            'result': True,
            'message': _('Diplomas generated.'),
            'names': names,
        })
